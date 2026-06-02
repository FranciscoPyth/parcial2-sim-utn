"""
Combina los 3 archivos .xlsx de cada parcial_turnoN en un único workbook
con una hoja por fragmento de iteraciones. Preserva valores, merges,
estilos básicos (font, fill, border, alignment, number_format) y dimensiones
de columnas/filas.
"""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\Usuario\General\Archivos Fran Dell\Ingeniería en Sistemas\4° Cuarto Año\Simulación (SIM)\parciales\recuperatorio_2")
OUT_DIR = Path(r"C:\Users\Usuario\General\Archivos Fran Dell\Ingeniería en Sistemas\4° Cuarto Año\Simulación (SIM)\landing-parcial2\assets")

TURNOS = {
    "Tema-1-Iteraciones.xlsx": [
        ("parcial_turno1/Filas 1 2 3.xlsx",     "Filas 1-3 (Cabecera)"),
        ("parcial_turno1/Filas 12 13 14.xlsx",  "Filas 12-14"),
        ("parcial_turno1/Filas 16 a 18.xlsx",   "Filas 16-18"),
    ],
    "Tema-2-Iteraciones.xlsx": [
        ("parcial_turno2/Filas 1 2 3 - 94654.xlsx",     "Filas 1-3 (Cabecera)"),
        ("parcial_turno2/Filas 11 12 13 - 94654.xlsx",  "Filas 11-13"),
        ("parcial_turno2/Filas 16 a 19 - 94654.xlsx",   "Filas 16-19"),
    ],
    "Tema-3-Iteraciones.xlsx": [
        ("parcial_turno3/Filas 1 2 3.xlsx", "Filas 1-3 (Cabecera)"),
        ("parcial_turno3/Filas 5 6 7.xlsx", "Filas 5-7"),
        ("parcial_turno3/Filas 16 a 19.xlsx", "Filas 16-19"),
    ],
}


def clear_student_rows(src_ws):
    """
    Vacía los valores de las filas que corresponden al trabajo del estudiante.
    Mantiene la PRIMERA fila con un Nro de Evento numérico (estado inicial del
    fragmento) y borra el contenido de las siguientes filas-evento.
    Preserva la cabecera, los merges, los estilos, y la sección auxiliar de
    RNDs (que aparece después, con texto en la columna A).
    """
    event_rows = []
    for r in range(1, src_ws.max_row + 1):
        v = src_ws.cell(row=r, column=1).value
        # Fila-evento = primera celda numérica entera (Nro Evento)
        if isinstance(v, (int, float)) and not isinstance(v, bool) and float(v).is_integer():
            event_rows.append(r)

    # Dejamos la primera (estado inicial) y borramos las siguientes
    for r in event_rows[1:]:
        for c in range(1, src_ws.max_column + 1):
            src_ws.cell(row=r, column=c).value = None


def copy_sheet(src_ws, dst_ws):
    # Merged ranges
    for mr in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mr))

    # Cell values + styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

    # Column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden

    # Row heights
    for r_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[r_idx].height = dim.height
        dst_ws.row_dimensions[r_idx].hidden = dim.hidden

    # Sheet view (freeze panes, zoom)
    dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes


def build(output_name, sources):
    wb_out = Workbook()
    # Remove default sheet
    wb_out.remove(wb_out.active)

    for rel_path, sheet_name in sources:
        src_file = BASE / rel_path
        wb_src = load_workbook(src_file)
        # Pick the first sheet (or the only relevant one)
        src_ws = wb_src.active  # active = first by default
        # If has multiple, pick the one with content
        for sn in wb_src.sheetnames:
            ws = wb_src[sn]
            if ws.max_row > 5:
                src_ws = ws
                break

        clear_student_rows(src_ws)
        dst_ws = wb_out.create_sheet(title=sheet_name[:31])  # Excel limit 31 chars
        copy_sheet(src_ws, dst_ws)

    out_path = OUT_DIR / output_name
    wb_out.save(out_path)
    print(f"OK: {out_path}")


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for out_name, sources in TURNOS.items():
        build(out_name, sources)
